"""
services/export_service.py
──────────────────────────
StoryForge Excel Export Service (V2)

Built for the current story format schema:
  linkedin / narrative / testimonial / case_study

Architecture:
  build_excel(stories, title)  →  io.BytesIO
      Takes a flat list of story dicts (from db_service),
      returns an in-memory Excel workbook ready for st.download_button.

  Sheets produced:
    • One sheet per format present in the data
      (LinkedIn Post / Long-Form Narrative / Testimonial / Case Study)
    • Summary sheet — counts, totals, export metadata

Intentionally simple:
  Readable, maintainable, ~200 lines total.
  No class hierarchy. No 500-line StyleEngine.
  openpyxl only.
"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.constants import STORY_FORMATS, AI_DISCLOSURE


# ── Design tokens ──────────────────────────────────────────────────────

_NAVY      = "0F172A"
_INDIGO    = "3B3F8C"
_TEAL      = "0D7377"
_SLATE     = "334155"
_WHITE     = "FFFFFF"
_OFF_WHITE = "F8FAFF"
_ROW_A     = "EEF2FF"
_ROW_B     = "FFFFFF"
_AMBER_BG  = "FFFBEB"
_AMBER_FG  = "B45309"
_GREEN_BG  = "F0FDF4"
_GREEN_FG  = "0D7377"
_RED_BG    = "FFF1F2"
_RED_FG    = "9B1C1C"
_MUTED     = "64748B"
_BODY      = "1E293B"
_BORDER    = "CBD5E1"

# Tab colours per format
_TAB_COLORS = {
    "linkedin":    "3B3F8C",
    "narrative":   "0D7377",
    "testimonial": "5B2D8E",
    "case_study":  "0F172A",
    "summary":     "334155",
}

# Status display
_STATUS_CFG = {
    "draft":     ("Draft",      _MUTED,    _ROW_B),
    "in_review": ("In Review",  _AMBER_FG, _AMBER_BG),
    "approved":  ("Approved",   _GREEN_FG, _GREEN_BG),
    "published": ("Published",  "5B2D8E",  "F5F3FF"),
    "rejected":  ("Rejected",   _RED_FG,   _RED_BG),
}


# ── Style helpers ──────────────────────────────────────────────────────

def _font(size=10, bold=False, color=_BODY, italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border():
    s = Side(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _header_cell(ws, row, col, text, bg=_NAVY):
    c = ws.cell(row, col, text)
    c.font      = _font(size=9, bold=True, color=_WHITE)
    c.fill      = _fill(bg)
    c.alignment = _align(h="center", wrap=True)
    c.border    = _border()
    return c


def _label_cell(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.font      = _font(size=9, bold=True)
    c.fill      = _fill(_ROW_A)
    c.alignment = _align(h="right")
    c.border    = _border()
    return c


def _value_cell(ws, row, col, text, bg=None):
    c = ws.cell(row, col, str(text) if text is not None else "—")
    c.font      = _font(size=9)
    c.fill      = _fill(bg or _ROW_B)
    c.alignment = _align(h="left", wrap=True)
    c.border    = _border()
    return c


def _banner(ws, row, text, bg, ncols, size=13):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=ncols,
    )
    c = ws.cell(row, 1, text)
    c.font      = _font(size=size, bold=True, color=_WHITE)
    c.fill      = _fill(bg)
    c.alignment = _align(h="center", v="center")
    c.border    = _border()
    ws.row_dimensions[row].height = 22


# ── Story sheet builder ────────────────────────────────────────────────

# Column layout for story sheets
_STORY_COLS = [
    ("Participant",  28),
    ("Program",      22),
    ("Domain",       18),
    ("Consent",      14),
    ("Status",       14),
    ("Word Count",   11),
    ("Story Content", 60),
    ("Editor Notes", 22),
    ("Generated On", 18),
]


def _build_story_sheet(ws, fmt_key, stories):
    """
    One worksheet for one story format.
    stories = list of story dicts (already filtered to this format).
    """
    spec  = STORY_FORMATS.get(fmt_key, {})
    label = spec.get("label", fmt_key)
    icon  = spec.get("icon", "")
    wmin, wmax = spec.get("word_range", (0, 9999))

    ws.sheet_properties.tabColor = _TAB_COLORS.get(fmt_key, _SLATE)
    ws.sheet_view.showGridLines  = False

    ncols = len(_STORY_COLS)

    # Title banner
    _banner(ws, 1, f"  {icon}  IAC StoryForge  \u00b7  {label}", _NAVY, ncols, size=13)
    _banner(ws, 2, f"{len(stories)} {'story' if len(stories)==1 else 'stories'}  \u00b7  "
                   f"Target: {wmin}\u2013{wmax} words", _INDIGO, ncols, size=9)

    # Column headers
    for ci, (h, _) in enumerate(_STORY_COLS, 1):
        _header_cell(ws, 3, ci, h, bg=_SLATE)

    # Data rows
    for ri, s in enumerate(stories, start=4):
        bg = _ROW_A if ri % 2 == 0 else _ROW_B

        # Status cell — colour-coded
        status     = s.get("status", "draft")
        s_label, s_fg, s_bg = _STATUS_CFG.get(status, ("Draft", _MUTED, _ROW_B))

        wc     = s.get("word_count", 0) or 0
        wc_ok  = wmin <= wc <= wmax
        wc_fg  = _GREEN_FG if wc_ok else _AMBER_FG
        wc_bg  = _GREEN_BG if wc_ok else _AMBER_BG

        created = s.get("created_at") or ""
        if created and "T" in created:
            created = created.split("T")[0]
        elif created and " " in created:
            created = created.split(" ")[0]

        row_data = [
            (s.get("participant_name", ""),       bg,   None),
            (s.get("program", "") or "—",         bg,   None),
            (s.get("domain", "") or "—",          bg,   None),
            (s.get("consent_level", "") or "—",   bg,   None),
            (s_label,                              s_bg, s_fg),
            (wc,                                   wc_bg, wc_fg),
            (s.get("content", "") or "",           _OFF_WHITE, None),
            (s.get("editor_notes", "") or "—",    bg,   None),
            (created,                              bg,   None),
        ]

        for ci, (val, cell_bg, cell_fg) in enumerate(row_data, 1):
            c = ws.cell(ri, ci, str(val) if val is not None else "—")
            c.fill   = _fill(cell_bg)
            c.border = _border()

            if ci == 6:  # Word count — centred mono
                c.font      = _font(size=9, bold=True, color=cell_fg or _BODY)
                c.alignment = _align(h="center")
            elif ci == 5:  # Status — centred bold
                c.font      = _font(size=9, bold=True, color=cell_fg or _BODY)
                c.alignment = _align(h="center")
            elif ci == 7:  # Story content — wrap, top-align
                c.font      = _font(size=9, color=_BODY)
                c.alignment = _align(h="left", v="top", wrap=True)
                ws.row_dimensions[ri].height = max(
                    ws.row_dimensions[ri].height or 15,
                    min(120, max(40, (wc // 12) * 5))
                )
            else:
                c.font      = _font(size=9, color=cell_fg or _BODY)
                c.alignment = _align(h="left", v="center", wrap=True)

    # AI Disclosure footer
    footer_row = 4 + len(stories) + 1
    ws.merge_cells(
        start_row=footer_row, start_column=1,
        end_row=footer_row,   end_column=ncols,
    )
    fc = ws.cell(footer_row, 1, AI_DISCLOSURE)
    fc.font      = _font(size=8, italic=True, color=_MUTED)
    fc.fill      = _fill("F1F5F9")
    fc.alignment = _align(h="left", v="center", wrap=True)
    fc.border    = _border()
    ws.row_dimensions[footer_row].height = 28

    # Column widths
    for ci, (_, width) in enumerate(_STORY_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = "A4"


# ── Summary sheet builder ──────────────────────────────────────────────

def _build_summary_sheet(ws, stories, title):
    ws.sheet_properties.tabColor = _TAB_COLORS["summary"]
    ws.sheet_view.showGridLines  = False
    NCOLS = 4

    now = datetime.now().strftime("%d %B %Y  %H:%M")

    _banner(ws, 1, "  IAC StoryForge  \u00b7  Export Summary", _NAVY, NCOLS, size=13)
    _banner(ws, 2, title, _INDIGO, NCOLS, size=9)

    # ── Section A: Totals ──────────────────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=NCOLS)
    sh = ws.cell(4, 1, "A.  EXPORT OVERVIEW")
    sh.font = _font(size=9, bold=True, color=_WHITE)
    sh.fill = _fill(_SLATE)
    sh.alignment = _align(h="center")
    sh.border = _border()

    overview = [
        ("Total Stories",      len(stories)),
        ("Total Participants", len({s.get("participant_id") for s in stories})),
        ("Total Words",        sum(s.get("word_count", 0) or 0 for s in stories)),
        ("Exported On",        now),
        ("Generator",          "StoryForge  \u00b7  Cloud Counselage / IAC"),
    ]
    r = 5
    for label, val in overview:
        _label_cell(ws, r, 1, label)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS)
        _value_cell(ws, r, 2, val, bg=_ROW_A if r % 2 == 0 else _ROW_B)
        r += 1

    # ── Section B: By Format ───────────────────────────────────────────
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    sh2 = ws.cell(r, 1, "B.  STORIES BY FORMAT")
    sh2.font = _font(size=9, bold=True, color=_WHITE)
    sh2.fill = _fill(_SLATE)
    sh2.alignment = _align(h="center")
    sh2.border = _border()
    r += 1

    for ci, h in enumerate(["Format", "Count", "Total Words", "Avg Words"], 1):
        _header_cell(ws, r, ci, h, bg=_INDIGO)
    r += 1

    for fmt_key, spec in STORY_FORMATS.items():
        fmt_stories = [s for s in stories if s.get("format") == fmt_key]
        if not fmt_stories:
            continue
        cnt   = len(fmt_stories)
        total = sum(s.get("word_count", 0) or 0 for s in fmt_stories)
        avg   = round(total / cnt) if cnt else 0
        bg    = _ROW_A if r % 2 == 0 else _ROW_B
        _value_cell(ws, r, 1, f"{spec['icon']} {spec['label']}", bg=bg)
        _value_cell(ws, r, 2, cnt, bg=bg)
        _value_cell(ws, r, 3, total, bg=bg)
        _value_cell(ws, r, 4, avg, bg=bg)
        r += 1

    # ── Section C: By Status ───────────────────────────────────────────
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    sh3 = ws.cell(r, 1, "C.  STORIES BY STATUS")
    sh3.font = _font(size=9, bold=True, color=_WHITE)
    sh3.fill = _fill(_SLATE)
    sh3.alignment = _align(h="center")
    sh3.border = _border()
    r += 1

    for ci, h in enumerate(["Status", "Count", "% of Total", ""], 1):
        _header_cell(ws, r, ci, h, bg=_INDIGO)
    r += 1

    total_cnt = len(stories) or 1
    for status_key, (status_label, s_fg, s_bg) in _STATUS_CFG.items():
        cnt = sum(1 for s in stories if s.get("status") == status_key)
        if cnt == 0:
            continue
        pct = f"{round(cnt / total_cnt * 100)}%"
        bg  = _ROW_A if r % 2 == 0 else _ROW_B
        c1 = ws.cell(r, 1, status_label)
        c1.font = _font(size=9, bold=True, color=s_fg)
        c1.fill = _fill(s_bg)
        c1.alignment = _align(h="center")
        c1.border = _border()
        _value_cell(ws, r, 2, cnt, bg=bg)
        _value_cell(ws, r, 3, pct, bg=bg)
        ws.cell(r, 4).fill   = _fill(bg)
        ws.cell(r, 4).border = _border()
        r += 1

    # Column widths
    for ci, w in enumerate([28, 14, 18, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════

def build_excel(stories: list, title: str = "IAC Story Export") -> io.BytesIO:
    """
    Build an Excel workbook from a list of story dicts and return it
    as an in-memory BytesIO object — ready for st.download_button.

    Each story dict must have at minimum:
        format, content, word_count, status, participant_name,
        program, domain (optional), consent_level (optional),
        editor_notes (optional), created_at (optional)

    Returns io.BytesIO (position reset to 0).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    # Summary sheet first
    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, stories, title)

    # One sheet per format (only formats present in the data)
    for fmt_key, spec in STORY_FORMATS.items():
        fmt_stories = [s for s in stories if s.get("format") == fmt_key]
        if not fmt_stories:
            continue
        ws = wb.create_sheet(spec["label"])
        _build_story_sheet(ws, fmt_key, fmt_stories)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_filename(label: str) -> str:
    """Generate a safe export filename with timestamp."""
    import re
    safe = re.sub(r"[^\w\-]", "_", label)
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    return f"IAC_StoryForge_{safe}_{ts}.xlsx"