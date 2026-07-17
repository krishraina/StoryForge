"""
services/excel_service.py
─────────────────────────
IAC Impact Story Generator — Professional Excel Reporting Service

ARCHITECTURE OVERVIEW:
──────────────────────
This module implements a three-layer reporting pipeline:

  1. ReportPayload   — Data layer.
                       Normalizes raw fields/stories dicts into a
                       validated, enriched object that ALL renderers
                       consume. Word counts, quality signals, timestamps,
                       and computed metadata live here — never scattered
                       across sheet builders.

  2. StyleEngine     — Primitive layer.
                       Every openpyxl styling call goes through this
                       class. Color changes, font size tweaks, border
                       styles — one place, nowhere else. Any future
                       renderer (PDF, DOCX) that needs styling constants
                       imports from here.

  3. SheetBuilder    — Transform layer.
                       Four dedicated methods, one per worksheet.
                       Each knows its audience and information hierarchy.
                       No sheet builder touches raw dicts — only payload.

  4. export_stories  — Public API.
                       The single function every caller uses: app.py,
                       batch upload, CLI scripts, future API endpoints.
                       Returns (True, filepath) or (False, error_message).

FUTURE EXTENSIBILITY:
─────────────────────
To add a PDF export in Phase 9:
    payload = ReportPayload(fields, stories)
    ok, path = PdfRenderer(payload).render()

To add DOCX:
    ok, path = DocxRenderer(payload).render()

To add a JSON API response:
    ok, data = JsonRenderer(payload).to_dict()

ReportPayload is the shared contract. Every renderer speaks the same
language. No renderer ever touches raw fields dicts again.

WORKSHEETS:
───────────
  Sheet 1 — Social Stories      Participant info + LinkedIn/IG story
  Sheet 2 — Grant Report        Participant info + formal funder story
  Sheet 3 — Newsletter          Participant info + community story
  Sheet 4 — Analytics Summary   Reporting dashboard: counts, quality,
                                 skills breakdown, approval tracker
"""

import os
import re
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from core.config import settings


# ══════════════════════════════════════════════════════════════════════
# DESIGN TOKENS
# Change the visual language of ALL exports here — nowhere else.
# ══════════════════════════════════════════════════════════════════════

class _T:
    """Theme constants — imported by StyleEngine, never directly."""

    # ── Primary palette ────────────────────────────────────
    DEEP_NAVY     = "0A1628"    # Darkest background / primary header
    NAVY          = "112240"    # Secondary header rows
    INDIGO        = "3B3F8C"    # Accent color — section titles
    INDIGO_LIGHT  = "5A5FC8"    # Sub-accent
    TEAL          = "0D7377"    # Success / approved / grant
    AMBER         = "B45309"    # Pending / warning
    ROSE          = "9B1C1C"    # Rejected / error
    PURPLE        = "5B2D8E"    # Newsletter accent
    SLATE         = "334155"    # Neutral dark
    STEEL         = "475569"    # Neutral mid

    # ── Text colors ────────────────────────────────────────
    WHITE         = "FFFFFF"
    OFF_WHITE     = "F1F5F9"
    BODY_DARK     = "1E293B"
    MUTED         = "64748B"

    # ── Row / cell backgrounds ─────────────────────────────
    ROW_EVEN      = "EEF2FF"    # Alternating row A
    ROW_ODD       = "FFFFFF"    # Alternating row B
    STORY_BG      = "F8FAFF"    # Story content cell
    META_BG       = "F0F4FF"    # Metadata cell
    SECTION_BG    = "E2E8F0"    # Section sub-header

    # ── Sheet tab colors ───────────────────────────────────
    TAB_SOCIAL    = "3B3F8C"
    TAB_GRANT     = "0D7377"
    TAB_NEWS      = "5B2D8E"
    TAB_ANALYTICS = "112240"

    # ── Border color ───────────────────────────────────────
    BORDER_CLR    = "CBD5E1"
    BORDER_DARK   = "334155"


# ══════════════════════════════════════════════════════════════════════
# STYLE ENGINE
# All openpyxl style primitives. Sheet builders call these methods
# instead of constructing styles inline.
# ══════════════════════════════════════════════════════════════════════

class StyleEngine:
    """
    Produces openpyxl style objects.

    Rules:
    - Every Font, Fill, Alignment, Border is built here.
    - SheetBuilder never imports openpyxl.styles directly.
    - Changing a style means changing one line in one method.
    """

    FONT_FACE = "Arial"

    @staticmethod
    def font(size=10, bold=False, color=_T.BODY_DARK,
             italic=False) -> Font:
        return Font(
            name=StyleEngine.FONT_FACE,
            size=size, bold=bold,
            color=color, italic=italic,
        )

    @staticmethod
    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def thin_border() -> Border:
        s = Side(style="thin", color=_T.BORDER_CLR)
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def thick_border() -> Border:
        tk = Side(style="medium", color=_T.BORDER_DARK)
        return Border(left=tk, right=tk, top=tk, bottom=tk)

    @staticmethod
    def align(h="left", v="center", wrap=False) -> Alignment:
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    @staticmethod
    def center(wrap=False) -> Alignment:
        return Alignment(horizontal="center", vertical="center",
                         wrap_text=wrap)

    # ── Compound cell stylers ──────────────────────────────

    @classmethod
    def apply_banner(cls, cell, text: str, bg: str,
                     size=13, bold=True):
        """Full-width banner: large bold text on solid color."""
        cell.value     = text
        cell.font      = cls.font(size=size, bold=bold, color=_T.WHITE)
        cell.fill      = cls.fill(bg)
        cell.alignment = cls.center(wrap=False)
        cell.border    = cls.thick_border()

    @classmethod
    def apply_section_header(cls, cell, text: str, bg: str,
                              size=10):
        """Section label row — slightly smaller, still bold."""
        cell.value     = text
        cell.font      = cls.font(size=size, bold=True, color=_T.WHITE)
        cell.fill      = cls.fill(bg)
        cell.alignment = cls.center(wrap=True)
        cell.border    = cls.thin_border()

    @classmethod
    def apply_col_header(cls, cell, text: str):
        """Column header inside a data table."""
        cell.value     = text
        cell.font      = cls.font(size=9, bold=True, color=_T.WHITE)
        cell.fill      = cls.fill(_T.NAVY)
        cell.alignment = cls.center(wrap=True)
        cell.border    = cls.thin_border()

    @classmethod
    def apply_label(cls, cell, text: str, row_idx: int):
        """Left-side field label in key-value pairs."""
        bg = _T.ROW_EVEN if row_idx % 2 == 0 else _T.ROW_ODD
        cell.value     = text
        cell.font      = cls.font(size=9, bold=True)
        cell.fill      = cls.fill(bg)
        cell.alignment = cls.align(h="right", v="center")
        cell.border    = cls.thin_border()

    @classmethod
    def apply_value(cls, cell, text: str, row_idx: int,
                    italic=False, color=_T.BODY_DARK):
        """Right-side field value in key-value pairs."""
        bg = _T.ROW_EVEN if row_idx % 2 == 0 else _T.ROW_ODD
        cell.value     = text
        cell.font      = cls.font(size=9, italic=italic, color=color)
        cell.fill      = cls.fill(bg)
        cell.alignment = cls.align(h="left", v="center", wrap=True)
        cell.border    = cls.thin_border()

    @classmethod
    def apply_story_cell(cls, cell, text: str, row_height_ref,
                         ws, row_num: int):
        """The main story content cell — wrapped, generous height."""
        cell.value     = text or "(Story not generated)"
        cell.font      = cls.font(size=10)
        cell.fill      = cls.fill(_T.STORY_BG)
        cell.alignment = cls.align(h="left", v="top", wrap=True)
        cell.border    = cls.thick_border()
        ws.row_dimensions[row_num].height = 180

    @classmethod
    def apply_status_cell(cls, cell, status: str):
        """Approval status cell with color coding."""
        color_map = {
            "approved":  (_T.TEAL,  "F0FDF4"),
            "pending":   (_T.AMBER, "FFFBEB"),
            "rejected":  (_T.ROSE,  "FFF1F2"),
        }
        key = status.lower()
        text_color, bg_color = color_map.get(key, (_T.MUTED, _T.ROW_ODD))
        cell.value     = {
            "approved": "✔  Approved",
            "pending":  "⏳  Pending Review",
            "rejected": "✘  Rejected",
        }.get(key, "⏳  Pending Review")
        cell.font      = cls.font(size=9, bold=True, color=text_color)
        cell.fill      = cls.fill(bg_color)
        cell.alignment = cls.center(wrap=False)
        cell.border    = cls.thin_border()

    @classmethod
    def apply_badge(cls, cell, value, good_range: tuple):
        """
        Word-count badge: green if within range, amber if outside.
        good_range = (low, high) inclusive.
        """
        low, high = good_range
        in_range   = low <= int(value or 0) <= high
        text_color = _T.TEAL if in_range else _T.AMBER
        bg_color   = "F0FDF4" if in_range else "FFFBEB"
        cell.value     = value
        cell.font      = cls.font(size=9, bold=True, color=text_color)
        cell.fill      = cls.fill(bg_color)
        cell.alignment = cls.center()
        cell.border    = cls.thin_border()


# ══════════════════════════════════════════════════════════════════════
# REPORT PAYLOAD
# The single source of truth for all export data.
# Normalizes raw input, computes all derived metrics, attaches metadata.
# Every SheetBuilder method reads from this object — never from dicts.
# ══════════════════════════════════════════════════════════════════════

class ReportPayload:
    """
    Validates and enriches raw story + participant data.

    Usage:
        payload = ReportPayload(fields, stories)
        # Then pass payload to SheetBuilder or any future renderer.

    Attributes exposed (all strings unless noted):
        Participant:  name, role, program, duration, location,
                      skills, achievement, outcome, impact, extra
        Stories:      social_story, grant_story, newsletter_story
        Computed:     wc_social, wc_grant, wc_newsletter  (int)
                      quality_social, quality_grant, quality_newsletter
                      stories_generated  (int: 0–3)
        Metadata:     timestamp, export_date, export_time,
                      ai_model, generator_version, organization
    """

    # Word-count target ranges per story type
    WC_TARGETS = {
        "social":     (80,  100),
        "grant":      (150, 170),
        "newsletter": (130, 150),
    }

    def __init__(self, fields: dict, stories: dict):
        # ── Participant fields ─────────────────────────────
        self.name        = (fields.get("name", "") or "").strip()
        self.role        = (fields.get("role", "") or "").strip()
        self.program     = (fields.get("program", "") or "").strip()
        self.duration    = (fields.get("duration", "") or "").strip()
        self.location    = (fields.get("location", "") or "").strip()
        self.skills      = (fields.get("skills", "") or "").strip()
        self.achievement = (fields.get("achievement", "") or "").strip()
        self.outcome     = (fields.get("outcome", "") or "").strip()
        self.impact      = (fields.get("impact", "") or "").strip()
        self.extra       = (fields.get("extra", "") or "").strip() or "—"

        # ── Stories ────────────────────────────────────────
        self.social_story     = (stories.get("social", "")     or "").strip()
        self.grant_story      = (stories.get("grant", "")      or "").strip()
        self.newsletter_story = (stories.get("newsletter", "") or "").strip()

        # ── Word counts (computed) ─────────────────────────
        self.wc_social     = self._wc(self.social_story)
        self.wc_grant      = self._wc(self.grant_story)
        self.wc_newsletter = self._wc(self.newsletter_story)
        self.wc_total      = self.wc_social + self.wc_grant + self.wc_newsletter

        # ── Quality signals (computed) ─────────────────────
        self.quality_social     = self._quality(self.social_story,     "social")
        self.quality_grant      = self._quality(self.grant_story,      "grant")
        self.quality_newsletter = self._quality(self.newsletter_story, "newsletter")

        # ── Story count ────────────────────────────────────
        self.stories_generated = sum([
            bool(self.social_story),
            bool(self.grant_story),
            bool(self.newsletter_story),
        ])

        # ── Skills list (computed) ─────────────────────────
        self.skills_list = [
            s.strip() for s in re.split(r"[,;/]", self.skills) if s.strip()
        ]

        # ── Metadata ───────────────────────────────────────
        now = datetime.now()
        self.timestamp         = now.strftime("%d %B %Y  %H:%M:%S")
        self.export_date       = now.strftime("%d %B %Y")
        self.export_time       = now.strftime("%H:%M:%S")
        self.ai_model          = "Google Gemini 2.0 Flash"
        self.generator_version = "IAC Story Generator v3.0"
        self.organization      = "Cloud Counselage Pvt. Ltd."
        self.initiative        = "Industry Academia Community (IAC) — Vision 2030"

    # ── Private helpers ────────────────────────────────────

    @staticmethod
    def _wc(text: str) -> int:
        return len(text.split()) if text and text.strip() else 0

    @classmethod
    def _quality(cls, text: str, story_type: str) -> str:
        """
        Lightweight heuristic quality signal.
        Not AI-based — checks structural indicators.
        Returns one of: "✔ On Target", "⚠ Check Length", "✘ Empty"
        """
        if not text or len(text.strip()) < 20:
            return "✘  Empty"
        wc        = cls._wc(text)
        low, high = cls.WC_TARGETS.get(story_type, (80, 200))
        if low <= wc <= high:
            return "✔  On Target"
        return f"⚠  {wc}w — target {low}–{high}"


# ══════════════════════════════════════════════════════════════════════
# SHEET BUILDER
# Four dedicated worksheet builders.
# Each method receives the worksheet (ws) and the payload.
# No raw dict access. No openpyxl style construction inline.
# ══════════════════════════════════════════════════════════════════════

class SheetBuilder:

    SE = StyleEngine   # Alias for brevity

    # ── Shared utilities ───────────────────────────────────

    @classmethod
    def _merge(cls, ws, r1, c1, r2, c2):
        ws.merge_cells(
            start_row=r1, start_column=c1,
            end_row=r2,   end_column=c2,
        )

    @classmethod
    def _banner_row(cls, ws, row: int, text: str, bg: str,
                    ncols: int, size=13):
        cls._merge(ws, row, 1, row, ncols)
        cls.SE.apply_banner(ws.cell(row, 1), text, bg, size=size)

    @classmethod
    def _section_row(cls, ws, row: int, text: str, bg: str,
                     ncols: int):
        cls._merge(ws, row, 1, row, ncols)
        cls.SE.apply_section_header(ws.cell(row, 1), text, bg)

    @classmethod
    def _kv_block(cls, ws, start_row: int, pairs: list,
                  value_span_to: int):
        """
        Renders a key-value block.
        pairs = [("Label", "value"), ...]
        value spans from column 2 to value_span_to.
        """
        for i, (label, value) in enumerate(pairs):
            row = start_row + i
            cls.SE.apply_label(ws.cell(row, 1), label, row)
            cls._merge(ws, row, 2, row, value_span_to)
            cls.SE.apply_value(ws.cell(row, 2), str(value), row)
        return start_row + len(pairs)

    @classmethod
    def _participant_block(cls, ws, payload: ReportPayload,
                           start_row: int, ncols: int,
                           accent: str) -> int:
        """
        Standard participant info block used by all 3 story sheets.
        Returns the next available row after the block.
        """
        cls._section_row(ws, start_row, "PARTICIPANT DETAILS",
                         accent, ncols)
        pairs = [
            ("Full Name",       payload.name),
            ("Role",            payload.role),
            ("Program",         payload.program),
            ("Duration",        payload.duration),
            ("Location",        payload.location),
            ("Impact Type",     payload.impact),
            ("Skills Gained",   payload.skills),
            ("Key Achievement", payload.achievement),
            ("Final Outcome",   payload.outcome),
            ("Extra Context",   payload.extra),
        ]
        return cls._kv_block(ws, start_row + 1, pairs, ncols)

    @classmethod
    def _metadata_block(cls, ws, payload: ReportPayload,
                        start_row: int, ncols: int,
                        accent: str) -> int:
        """
        Standard export metadata block used by all 3 story sheets.
        Returns the next available row.
        """
        cls._section_row(ws, start_row, "EXPORT METADATA",
                         _T.SLATE, ncols)
        pairs = [
            ("Generated On",    payload.timestamp),
            ("AI Model",        payload.ai_model),
            ("Generator",       payload.generator_version),
            ("Organization",    payload.organization),
            ("Initiative",      payload.initiative),
        ]
        return cls._kv_block(ws, start_row + 1, pairs, ncols)

    # ── Sheet 1: Social Media Stories ─────────────────────

    @classmethod
    def build_social_sheet(cls, ws, payload: ReportPayload):
        """
        Audience: Social media managers, participant themselves.
        Priority: Story quality, word count compliance, hashtag presence.
        """
        ws.sheet_properties.tabColor = _T.TAB_SOCIAL
        ws.sheet_view.showGridLines   = False
        NCOLS = 4

        # ── Title banner ──────────────────────────────────
        cls._banner_row(ws, 1,
            "✦  IAC IMPACT STORY — SOCIAL MEDIA",
            _T.DEEP_NAVY, NCOLS, size=14)
        cls._banner_row(ws, 2,
            "LinkedIn / Instagram  ·  First Person  ·  80–100 Words",
            _T.INDIGO, NCOLS, size=10)

        # ── Participant block ──────────────────────────────
        next_row = cls._participant_block(
            ws, payload, 4, NCOLS, _T.INDIGO)

        # ── Story section ──────────────────────────────────
        next_row += 1
        cls._section_row(ws, next_row,
            "GENERATED STORY — READY FOR PUBLISHING",
            _T.INDIGO, NCOLS)
        next_row += 1

        # Story + status side-by-side
        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=3)
        cls.SE.apply_story_cell(
            ws.cell(next_row, 1),
            payload.social_story, None, ws, next_row)

        cls.SE.apply_status_cell(ws.cell(next_row, 4), "pending")

        # ── Story metrics row ──────────────────────────────
        next_row += 1
        metrics = [
            ("Word Count",    payload.wc_social,      (80, 100)),
            ("Target",        "80 – 100 words",       None),
            ("Quality",       payload.quality_social,  None),
            ("Approval",      "⏳  Pending Review",   None),
        ]
        for col_i, (label, value, wc_range) in enumerate(metrics, 1):
            lc = ws.cell(next_row, col_i)
            cls.SE.apply_section_header(lc, label, _T.SLATE)

            vc = ws.cell(next_row + 1, col_i)
            if wc_range:
                cls.SE.apply_badge(vc, value, wc_range)
            else:
                vc.value     = str(value)
                vc.font      = cls.SE.font(size=9, bold=True)
                vc.fill      = cls.SE.fill(_T.SECTION_BG)
                vc.alignment = cls.SE.center()
                vc.border    = cls.SE.thin_border()

        # ── Usage guidance ─────────────────────────────────
        next_row += 3
        cls._section_row(ws, next_row, "PUBLISHING GUIDANCE",
                         _T.SLATE, NCOLS)
        next_row += 1
        guidance_text = (
            "• Post on LinkedIn on Tuesday or Wednesday between 8–10am for maximum reach.\n"
            "• Attach a clear, professional photo — engagement increases 3× with imagery.\n"
            "• Tag @CloudCounselage and @IAC in your post for wider distribution.\n"
            "• Use all hashtags exactly as generated — they are research-optimised.\n"
            "• Stories work well as Instagram carousels: one slide per paragraph."
        )
        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=NCOLS)
        gc = ws.cell(next_row, 1)
        gc.value     = guidance_text
        gc.font      = cls.SE.font(size=9, italic=True)
        gc.fill      = cls.SE.fill(_T.META_BG)
        gc.alignment = cls.SE.align(h="left", v="top", wrap=True)
        gc.border    = cls.SE.thin_border()
        ws.row_dimensions[next_row].height = 70

        # ── Metadata ───────────────────────────────────────
        next_row += 2
        cls._metadata_block(ws, payload, next_row, NCOLS, _T.INDIGO)

        # ── Column widths ──────────────────────────────────
        for col, width in zip([1, 2, 3, 4], [60, 18, 20, 22]):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A4"

    # ── Sheet 2: Grant Report Stories ─────────────────────

    @classmethod
    def build_grant_sheet(cls, ws, payload: ReportPayload):
        """
        Audience: NGO funders, government partners, CSR teams.
        Priority: Formal tone, measurable outcomes, Vision 2030 alignment.
        """
        ws.sheet_properties.tabColor = _T.TAB_GRANT
        ws.sheet_view.showGridLines   = False
        NCOLS = 4

        cls._banner_row(ws, 1,
            "✦  IAC IMPACT STORY — GRANT REPORT",
            _T.DEEP_NAVY, NCOLS, size=14)
        cls._banner_row(ws, 2,
            "Formal Third Person  ·  Funder Ready  ·  150–170 Words",
            _T.TEAL, NCOLS, size=10)

        next_row = cls._participant_block(
            ws, payload, 4, NCOLS, _T.TEAL)

        next_row += 1
        cls._section_row(ws, next_row,
            "GENERATED STORY — SUITABLE FOR GRANT APPLICATIONS & ANNUAL REPORTS",
            _T.TEAL, NCOLS)
        next_row += 1

        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=3)
        cls.SE.apply_story_cell(
            ws.cell(next_row, 1),
            payload.grant_story, None, ws, next_row)
        cls.SE.apply_status_cell(ws.cell(next_row, 4), "pending")

        next_row += 1
        metrics = [
            ("Word Count",  payload.wc_grant,      (150, 170)),
            ("Target",      "150 – 170 words",      None),
            ("Quality",     payload.quality_grant,   None),
            ("Approval",    "⏳  Pending Review",   None),
        ]
        for col_i, (label, value, wc_range) in enumerate(metrics, 1):
            cls.SE.apply_section_header(ws.cell(next_row, col_i),
                                        label, _T.SLATE)
            vc = ws.cell(next_row + 1, col_i)
            if wc_range:
                cls.SE.apply_badge(vc, value, wc_range)
            else:
                vc.value     = str(value)
                vc.font      = cls.SE.font(size=9, bold=True)
                vc.fill      = cls.SE.fill(_T.SECTION_BG)
                vc.alignment = cls.SE.center()
                vc.border    = cls.SE.thin_border()

        next_row += 3
        cls._section_row(ws, next_row, "USAGE GUIDANCE",
                         _T.SLATE, NCOLS)
        next_row += 1
        guidance_text = (
            "• Suitable for inclusion in NGO annual impact reports as-is.\n"
            "• Pair with participant photograph and outcome data for NASSCOM / DST grant applications.\n"
            "• Reference IAC's Vision 2030 and Cloud Counselage's mission when presenting to government funders.\n"
            "• This format meets MeitY CSR documentation standards for skill-development reporting.\n"
            "• Add quantitative program metrics (total participants, placement rate) alongside this story."
        )
        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=NCOLS)
        gc = ws.cell(next_row, 1)
        gc.value     = guidance_text
        gc.font      = cls.SE.font(size=9, italic=True)
        gc.fill      = cls.SE.fill(_T.META_BG)
        gc.alignment = cls.SE.align(h="left", v="top", wrap=True)
        gc.border    = cls.SE.thin_border()
        ws.row_dimensions[next_row].height = 75

        next_row += 2
        cls._metadata_block(ws, payload, next_row, NCOLS, _T.TEAL)

        for col, width in zip([1, 2, 3, 4], [65, 18, 20, 22]):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A4"

    # ── Sheet 3: Newsletter Stories ────────────────────────

    @classmethod
    def build_newsletter_sheet(cls, ws, payload: ReportPayload):
        """
        Audience: IAC community members, prospective participants.
        Priority: Warmth, relatability, call-to-action for future joiners.
        """
        ws.sheet_properties.tabColor = _T.TAB_NEWS
        ws.sheet_view.showGridLines   = False
        NCOLS = 4

        cls._banner_row(ws, 1,
            "✦  IAC IMPACT STORY — NEWSLETTER",
            _T.DEEP_NAVY, NCOLS, size=14)
        cls._banner_row(ws, 2,
            "First Person  ·  Community Tone  ·  130–150 Words",
            _T.PURPLE, NCOLS, size=10)

        next_row = cls._participant_block(
            ws, payload, 4, NCOLS, _T.PURPLE)

        next_row += 1
        cls._section_row(ws, next_row,
            "GENERATED STORY — READY FOR IAC MONTHLY NEWSLETTER",
            _T.PURPLE, NCOLS)
        next_row += 1

        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=3)
        cls.SE.apply_story_cell(
            ws.cell(next_row, 1),
            payload.newsletter_story, None, ws, next_row)
        cls.SE.apply_status_cell(ws.cell(next_row, 4), "pending")

        next_row += 1
        metrics = [
            ("Word Count", payload.wc_newsletter, (130, 150)),
            ("Target",     "130 – 150 words",     None),
            ("Quality",    payload.quality_newsletter, None),
            ("Approval",   "⏳  Pending Review",  None),
        ]
        for col_i, (label, value, wc_range) in enumerate(metrics, 1):
            cls.SE.apply_section_header(ws.cell(next_row, col_i),
                                        label, _T.SLATE)
            vc = ws.cell(next_row + 1, col_i)
            if wc_range:
                cls.SE.apply_badge(vc, value, wc_range)
            else:
                vc.value     = str(value)
                vc.font      = cls.SE.font(size=9, bold=True)
                vc.fill      = cls.SE.fill(_T.SECTION_BG)
                vc.alignment = cls.SE.center()
                vc.border    = cls.SE.thin_border()

        next_row += 3
        cls._section_row(ws, next_row, "EDITORIAL GUIDANCE",
                         _T.SLATE, NCOLS)
        next_row += 1
        guidance_text = (
            "• Feature in the top section of the newsletter — above-the-fold placement increases read-through by 40%.\n"
            "• Add a 1-line bio and headshot below the story for authenticity.\n"
            "• Include a call-to-action: 'Apply to IAC at cloudcounselage.com/apply'\n"
            "• This story works well as a standalone WhatsApp / Telegram community post.\n"
            "• Consider pairing with a 30-second video testimonial from the same participant."
        )
        ws.merge_cells(
            start_row=next_row, start_column=1,
            end_row=next_row,   end_column=NCOLS)
        gc = ws.cell(next_row, 1)
        gc.value     = guidance_text
        gc.font      = cls.SE.font(size=9, italic=True)
        gc.fill      = cls.SE.fill(_T.META_BG)
        gc.alignment = cls.SE.align(h="left", v="top", wrap=True)
        gc.border    = cls.SE.thin_border()
        ws.row_dimensions[next_row].height = 75

        next_row += 2
        cls._metadata_block(ws, payload, next_row, NCOLS, _T.PURPLE)

        for col, width in zip([1, 2, 3, 4], [62, 18, 20, 22]):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A4"

    # ── Sheet 4: Analytics Summary ─────────────────────────

    @classmethod
    def build_analytics_sheet(cls, ws, payload: ReportPayload):
        """
        Audience: Program managers, internship coordinators, data reviewers.
        Priority: At-a-glance metrics, quality signals, approval workflow.

        This sheet is designed to scale: when batch upload (Phase 7)
        adds multiple participants, this sheet will aggregate across all
        of them. The row structure is already built for that.
        """
        ws.sheet_properties.tabColor = _T.TAB_ANALYTICS
        ws.sheet_view.showGridLines   = False
        NCOLS = 5

        # ── Title ──────────────────────────────────────────
        cls._banner_row(ws, 1,
            "✦  IAC ANALYTICS & REPORTING DASHBOARD",
            _T.DEEP_NAVY, NCOLS, size=14)
        cls._banner_row(ws, 2,
            "Story Quality Metrics  ·  Word Count Analysis  ·  Approval Tracker",
            _T.NAVY, NCOLS, size=10)

        # ── Section A: Generation Summary ─────────────────
        cls._section_row(ws, 4,
            "A.  GENERATION SUMMARY", _T.INDIGO, NCOLS)

        summary_pairs = [
            ("Participant Name",   payload.name),
            ("Role",               payload.role),
            ("Program",            payload.program),
            ("Duration",           payload.duration),
            ("Location",           payload.location),
            ("Impact Category",    payload.impact),
            ("Stories Generated",  f"{payload.stories_generated} / 3"),
            ("Total Word Count",   f"{payload.wc_total} words across all stories"),
            ("AI Model Used",      payload.ai_model),
            ("Generated On",       payload.timestamp),
            ("Exported By",        payload.generator_version),
        ]
        next_row = 5
        for i, (label, value) in enumerate(summary_pairs):
            row = next_row + i
            cls.SE.apply_label(ws.cell(row, 1), label, row)
            ws.merge_cells(
                start_row=row, start_column=2,
                end_row=row,   end_column=NCOLS)
            cls.SE.apply_value(ws.cell(row, 2), str(value), row)

        # ── Section B: Story Quality Metrics Table ─────────
        b_start = next_row + len(summary_pairs) + 2
        cls._section_row(ws, b_start,
            "B.  STORY QUALITY METRICS", _T.INDIGO, NCOLS)

        b_header = b_start + 1
        col_headers = [
            "Story Type", "Word Count",
            "Target Range", "Compliance", "Quality Signal",
        ]
        for ci, h in enumerate(col_headers, 1):
            cls.SE.apply_col_header(ws.cell(b_header, ci), h)

        quality_data = [
            ("Social Media",  payload.wc_social,
             "80–100",  payload.quality_social,
             (80, 100)),
            ("Grant Report",  payload.wc_grant,
             "150–170", payload.quality_grant,
             (150, 170)),
            ("Newsletter",    payload.wc_newsletter,
             "130–150", payload.quality_newsletter,
             (130, 150)),
        ]
        for qi, (label, wc, target, quality, wc_range) in \
                enumerate(quality_data, start=b_header + 1):
            bg = _T.ROW_EVEN if qi % 2 == 0 else _T.ROW_ODD

            ws.cell(qi, 1).value     = label
            ws.cell(qi, 1).font      = cls.SE.font(size=9, bold=True)
            ws.cell(qi, 1).fill      = cls.SE.fill(bg)
            ws.cell(qi, 1).alignment = cls.SE.center()
            ws.cell(qi, 1).border    = cls.SE.thin_border()

            cls.SE.apply_badge(ws.cell(qi, 2), wc, wc_range)

            ws.cell(qi, 3).value     = target
            ws.cell(qi, 3).font      = cls.SE.font(size=9)
            ws.cell(qi, 3).fill      = cls.SE.fill(bg)
            ws.cell(qi, 3).alignment = cls.SE.center()
            ws.cell(qi, 3).border    = cls.SE.thin_border()

            # Compliance column
            in_range   = wc_range[0] <= wc <= wc_range[1]
            comp_text  = "✔  Within Target" if in_range else "⚠  Out of Range"
            comp_color = _T.TEAL if in_range else _T.AMBER
            comp_bg    = "F0FDF4" if in_range else "FFFBEB"
            ws.cell(qi, 4).value     = comp_text
            ws.cell(qi, 4).font      = cls.SE.font(size=9, bold=True,
                                                    color=comp_color)
            ws.cell(qi, 4).fill      = cls.SE.fill(comp_bg)
            ws.cell(qi, 4).alignment = cls.SE.center()
            ws.cell(qi, 4).border    = cls.SE.thin_border()

            ws.cell(qi, 5).value     = quality
            ws.cell(qi, 5).font      = cls.SE.font(size=9)
            ws.cell(qi, 5).fill      = cls.SE.fill(bg)
            ws.cell(qi, 5).alignment = cls.SE.center()
            ws.cell(qi, 5).border    = cls.SE.thin_border()

        # ── Section C: Skills Breakdown ────────────────────
        c_start = b_header + len(quality_data) + 3
        cls._section_row(ws, c_start,
            "C.  SKILLS PROFILE", _T.SLATE, NCOLS)

        if payload.skills_list:
            for si, skill in enumerate(payload.skills_list):
                row = c_start + 1 + si
                bg  = _T.ROW_EVEN if si % 2 == 0 else _T.ROW_ODD

                ws.cell(row, 1).value     = f"Skill {si + 1}"
                ws.cell(row, 1).font      = cls.SE.font(size=9, bold=True)
                ws.cell(row, 1).fill      = cls.SE.fill(bg)
                ws.cell(row, 1).alignment = cls.SE.align(h="right", v="center")
                ws.cell(row, 1).border    = cls.SE.thin_border()

                ws.merge_cells(
                    start_row=row, start_column=2,
                    end_row=row,   end_column=NCOLS)
                ws.cell(row, 2).value     = skill
                ws.cell(row, 2).font      = cls.SE.font(size=9)
                ws.cell(row, 2).fill      = cls.SE.fill(bg)
                ws.cell(row, 2).alignment = cls.SE.align(h="left", v="center")
                ws.cell(row, 2).border    = cls.SE.thin_border()
            skills_end = c_start + 1 + len(payload.skills_list)
        else:
            skills_end = c_start + 2

        # ── Section D: Approval Tracker ────────────────────
        d_start = skills_end + 2
        cls._section_row(ws, d_start,
            "D.  APPROVAL WORKFLOW TRACKER", _T.INDIGO, NCOLS)

        d_header = d_start + 1
        approval_cols = [
            "Story Type", "Assigned To",
            "Current Status", "Review Date", "Final Decision",
        ]
        for ci, h in enumerate(approval_cols, 1):
            cls.SE.apply_col_header(ws.cell(d_header, ci), h)

        approval_data = [
            ("Social Media Story",  "Content / Marketing Team",  "⏳  Pending"),
            ("Grant Report Story",  "Program Manager",           "⏳  Pending"),
            ("Newsletter Story",    "Editorial / Comms Team",    "⏳  Pending"),
        ]
        for ai, (story_type, assignee, status) in \
                enumerate(approval_data, start=d_header + 1):
            bg = _T.ROW_EVEN if ai % 2 == 0 else _T.ROW_ODD
            row_data = [story_type, assignee, status, "", ""]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(ai, ci)
                cell.value     = val
                cell.font      = cls.SE.font(
                    size=9,
                    bold=(ci == 1),
                    color=_T.AMBER if ci == 3 else _T.BODY_DARK,
                )
                cell.fill      = cls.SE.fill(bg)
                cell.alignment = cls.SE.center(wrap=False)
                cell.border    = cls.SE.thin_border()

        # ── Section E: Export Statistics ───────────────────
        e_start = d_header + len(approval_data) + 3
        cls._section_row(ws, e_start,
            "E.  EXPORT STATISTICS", _T.SLATE, NCOLS)

        export_pairs = [
            ("Export Format",    "Microsoft Excel Workbook (.xlsx)"),
            ("Total Worksheets", "4  (Social Media, Grant Report, Newsletter, Analytics)"),
            ("Export Timestamp", payload.timestamp),
            ("Generator",        payload.generator_version),
            ("Powered By",       payload.ai_model),
            ("Organization",     payload.organization),
            ("Initiative",       payload.initiative),
        ]
        for ei, (label, value) in enumerate(export_pairs):
            row = e_start + 1 + ei
            cls.SE.apply_label(ws.cell(row, 1), label, row)
            ws.merge_cells(
                start_row=row, start_column=2,
                end_row=row,   end_column=NCOLS)
            cls.SE.apply_value(ws.cell(row, 2), value, row)

        # ── Column widths ──────────────────────────────────
        for col, width in zip([1, 2, 3, 4, 5], [28, 30, 20, 18, 18]):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════
# PUBLIC API
# This is the ONLY function any other module ever calls.
# ══════════════════════════════════════════════════════════════════════

def export_stories(
    fields: dict,
    stories: dict,
    custom_filename: str = None,
) -> tuple[bool, str]:
    """
    Generate a production-grade Excel report for one participant.

    Args:
        fields          : Participant details dict (name, role, skills…)
        stories         : {"social": "…", "grant": "…", "newsletter": "…"}
        custom_filename : Optional filename. Auto-generated if omitted.

    Returns:
        (True,  "/abs/path/to/file.xlsx")  on success
        (False, "error description")        on failure

    This function is intentionally side-effect-free beyond writing
    the Excel file. It does not touch the database, Streamlit session,
    or Gemini. Any caller — UI, batch script, API endpoint — gets
    the same consistent result.
    """
    try:
        # ── 1. Build the payload (validates + enriches data) ───
        payload = ReportPayload(fields, stories)

        # ── 2. Create workbook ─────────────────────────────────
        wb = openpyxl.Workbook()
        wb.remove(wb.active)          # Remove default empty sheet

        # ── 3. Add worksheets ──────────────────────────────────
        ws_social   = wb.create_sheet("Social Media")
        ws_grant    = wb.create_sheet("Grant Report")
        ws_news     = wb.create_sheet("Newsletter")
        ws_analytics = wb.create_sheet("Analytics")

        # ── 4. Delegate to SheetBuilder ────────────────────────
        SheetBuilder.build_social_sheet(ws_social, payload)
        SheetBuilder.build_grant_sheet(ws_grant, payload)
        SheetBuilder.build_newsletter_sheet(ws_news, payload)
        SheetBuilder.build_analytics_sheet(ws_analytics, payload)

        # ── 5. Build filename ──────────────────────────────────
        if custom_filename:
            filename = custom_filename
        else:
            safe_name = re.sub(r"[^\w\-]", "_", payload.name or "Participant")
            timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
            filename   = f"IAC_Story_{safe_name}_{timestamp}.xlsx"

        # ── 6. Save ────────────────────────────────────────────
        os.makedirs(settings.OUTPUT_DIR, exist_ok=True)
        filepath = os.path.join(settings.OUTPUT_DIR, filename)
        wb.save(filepath)

        return True, filepath

    except Exception as exc:
        return False, f"Export failed: {exc}"